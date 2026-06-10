"""考试管理路由"""
import io
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.models import Exam, Score, ExamAnswer, ExamRegistration, Student
from app.websocket_manager import manager as ws_manager
from app.schemas import (
    ExamCreate, ExamUpdate, ExamResponse,
    ExamWithStats, ExamRegistrationCreate,
    ExamRegistrationUpdate, ExamRegistrationResponse,
    ExamRegistrationDetail, PaginatedResponse
)

router = APIRouter(prefix="/api/exams", tags=["考试管理"])


# ==================== 考试 CRUD ====================

@router.get("", summary="获取考试列表", response_model=PaginatedResponse)
def get_exams(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    keyword: Optional[str] = Query(None, description="搜索关键词(考试名称)"),
    subject: Optional[str] = Query(None, description="科目筛选"),
    status: Optional[str] = Query(None, description="状态筛选"),
    exam_type: Optional[str] = Query(None, description="类型筛选"),
    db: Session = Depends(get_db)
):
    """分页获取考试列表，支持搜索和筛选"""
    query = db.query(Exam)

    if keyword:
        query = query.filter(Exam.name.like(f"%{keyword}%"))
    if subject:
        query = query.filter(Exam.subject.like(f"%{subject}%"))
    if status:
        query = query.filter(Exam.status == status)
    if exam_type:
        query = query.filter(Exam.exam_type == exam_type)

    total = query.count()
    total_pages = (total + page_size - 1) // page_size

    exams = query.order_by(Exam.exam_date.desc()) \
        .offset((page - 1) * page_size) \
        .limit(page_size) \
        .all()

    items = []
    for e in exams:
        registered_count = db.query(func.count(ExamRegistration.id)).filter(
            ExamRegistration.exam_id == e.id
        ).scalar() or 0

        scored_count = db.query(func.count(Score.id)).filter(
            Score.exam_id == e.id
        ).scalar() or 0

        score_stats = db.query(
            func.avg(Score.score).label("avg"),
            func.max(Score.score).label("highest"),
            func.min(Score.score).label("lowest"),
        ).filter(Score.exam_id == e.id).first()

        pass_count = db.query(func.count(Score.id)).filter(
            Score.exam_id == e.id,
            Score.passed == True
        ).scalar() or 0

        avg_score = round(score_stats.avg, 2) if score_stats.avg else None
        highest_score = round(score_stats.highest, 2) if score_stats.highest else None
        lowest_score = round(score_stats.lowest, 2) if score_stats.lowest else None
        pass_rate = round(pass_count / scored_count * 100, 2) if scored_count > 0 else None

        items.append(ExamWithStats(
            id=e.id,
            name=e.name,
            subject=e.subject,
            exam_type=e.exam_type,
            exam_date=e.exam_date,
            duration=e.duration,
            location=e.location,
            total_score=e.total_score,
            pass_score=e.pass_score,
            status=e.status,
            description=e.description,
            max_participants=e.max_participants,
            created_at=e.created_at,
            updated_at=e.updated_at,
            registered_count=registered_count,
            scored_count=scored_count,
            avg_score=avg_score,
            pass_rate=pass_rate,
            highest_score=highest_score,
            lowest_score=lowest_score,
        ))

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@router.get("/upcoming", summary="获取即将开始的考试")
def get_upcoming_exams(
    limit: int = Query(5, ge=1, le=20, description="返回数量"),
    db: Session = Depends(get_db)
):
    """获取即将开始的考试列表"""
    now = datetime.now()
    exams = db.query(Exam).filter(
        Exam.exam_date > now,
        Exam.status.in_(["upcoming", "ongoing"])
    ).order_by(Exam.exam_date.asc()).limit(limit).all()

    return {
        "code": 200,
        "message": "success",
        "data": [
            {
                "id": e.id,
                "name": e.name,
                "subject": e.subject,
                "exam_type": e.exam_type,
                "exam_date": str(e.exam_date),
                "duration": e.duration,
                "location": e.location,
                "status": e.status,
            }
            for e in exams
        ]
    }


@router.get("/{exam_id}", summary="获取考试详情", response_model=ExamWithStats)
def get_exam(exam_id: int, db: Session = Depends(get_db)):
    """根据ID获取考试详情"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    registered_count = db.query(func.count(ExamRegistration.id)).filter(
        ExamRegistration.exam_id == exam.id
    ).scalar() or 0

    scored_count = db.query(func.count(Score.id)).filter(
        Score.exam_id == exam.id
    ).scalar() or 0

    score_stats = db.query(
        func.avg(Score.score).label("avg"),
        func.max(Score.score).label("highest"),
        func.min(Score.score).label("lowest"),
    ).filter(Score.exam_id == exam.id).first()

    pass_count = db.query(func.count(Score.id)).filter(
        Score.exam_id == exam.id,
        Score.passed == True
    ).scalar() or 0

    avg_score = round(score_stats.avg, 2) if score_stats.avg else None
    highest_score = round(score_stats.highest, 2) if score_stats.highest else None
    lowest_score = round(score_stats.lowest, 2) if score_stats.lowest else None
    pass_rate = round(pass_count / scored_count * 100, 2) if scored_count > 0 else None

    return ExamWithStats(
        id=exam.id,
        name=exam.name,
        subject=exam.subject,
        exam_type=exam.exam_type,
        exam_date=exam.exam_date,
        duration=exam.duration,
        location=exam.location,
        total_score=exam.total_score,
        pass_score=exam.pass_score,
        status=exam.status,
        description=exam.description,
        max_participants=exam.max_participants,
        created_at=exam.created_at,
        updated_at=exam.updated_at,
        registered_count=registered_count,
        scored_count=scored_count,
        avg_score=avg_score,
        pass_rate=pass_rate,
        highest_score=highest_score,
        lowest_score=lowest_score,
    )


@router.post("", summary="创建考试", response_model=ExamResponse)
def create_exam(exam_data: ExamCreate, db: Session = Depends(get_db)):
    """创建新考试"""
    exam = Exam(**exam_data.model_dump())
    db.add(exam)
    db.commit()
    db.refresh(exam)
    return exam


@router.put("/{exam_id}", summary="更新考试信息", response_model=ExamResponse)
def update_exam(exam_id: int, exam_data: ExamUpdate, db: Session = Depends(get_db)):
    """更新考试信息"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    update_data = exam_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(exam, key, value)

    db.commit()
    db.refresh(exam)
    return exam


@router.delete("/{exam_id}", summary="删除考试")
def delete_exam(exam_id: int, db: Session = Depends(get_db)):
    """删除考试"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    db.delete(exam)
    db.commit()
    return {"code": 200, "message": "删除成功"}


@router.post("/{exam_id}/start", summary="开始考试")
def start_exam(exam_id: int, db: Session = Depends(get_db)):
    """将考试状态改为进行中"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    if exam.status not in ("upcoming", "cancelled"):
        raise HTTPException(status_code=400, detail=f"当前状态「{exam.status}」无法开始考试，只有「即将开始」或「已取消」状态的考试可以开始")

    exam.status = "ongoing"
    db.commit()
    db.refresh(exam)
    return {"code": 200, "message": "考试已开始", "data": {"id": exam.id, "status": exam.status}}


@router.post("/{exam_id}/end", summary="结束考试")
async def end_exam(exam_id: int, db: Session = Depends(get_db)):
    """将考试状态改为已结束，并通过 WebSocket 实时通知学生端"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    if exam.status != "ongoing":
        raise HTTPException(status_code=400, detail=f"当前状态「{exam.status}」无法结束考试，只有「进行中」状态的考试可以结束")

    exam.status = "finished"
    db.commit()
    db.refresh(exam)

    # 通过 WebSocket 立即通知所有正在考试的学生
    await ws_manager.broadcast_to_exam(exam_id, {
        "type": "exam_ended",
        "reason": "admin_end",
        "message": "管理员已结束本次考试"
    })

    return {"code": 200, "message": "考试已结束", "data": {"id": exam.id, "status": exam.status}}


@router.post("/{exam_id}/force-end", summary="强制结束考试（管理端修改题目后）")
async def force_end_exam(exam_id: int, db: Session = Depends(get_db)):
    """强制结束考试：清除所有成绩和答题记录，将考试状态改为已结束，并通过 WebSocket 实时通知学生端"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    # 删除该考试所有学员的答题记录
    db.query(ExamAnswer).filter(ExamAnswer.exam_id == exam_id).delete()

    # 删除该考试所有学员的成绩记录
    db.query(Score).filter(Score.exam_id == exam_id).delete()

    # 将考试状态设为已结束
    exam.status = "finished"
    db.commit()
    db.refresh(exam)

    # 通过 WebSocket 立即通知所有正在考试的学生
    await ws_manager.broadcast_to_exam(exam_id, {
        "type": "exam_ended",
        "reason": "force_end",
        "message": "管理员已修改题目，考试已强制结束"
    })

    return {"code": 200, "message": f"考试「{exam.name}」已强制结束，所有成绩和答题记录已清除", "data": {"id": exam.id, "status": exam.status}}


@router.post("/{exam_id}/retake", summary="重新考试")
def retake_exam(exam_id: int, db: Session = Depends(get_db)):
    """重置考试：清除所有成绩和答题记录，将考试状态改为进行中，允许学员重新参加考试"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    if exam.status != "finished":
        raise HTTPException(status_code=400, detail=f"当前状态「{exam.status}」无法重新考试，只有「已结束」状态的考试可以重新开始")

    # 删除该考试所有学员的答题记录
    db.query(ExamAnswer).filter(ExamAnswer.exam_id == exam_id).delete()

    # 删除该考试所有学员的成绩记录
    db.query(Score).filter(Score.exam_id == exam_id).delete()

    # 将考试状态改回进行中
    exam.status = "ongoing"
    db.commit()
    db.refresh(exam)
    return {"code": 200, "message": f"考试「{exam.name}」已重置，所有成绩和答题记录已清除，学员可重新参加考试", "data": {"id": exam.id, "status": exam.status}}


# ==================== 考试报名管理 ====================

@router.post("/{exam_id}/register", summary="学员报名考试", response_model=ExamRegistrationResponse)
def register_exam(
    exam_id: int,
    registration: ExamRegistrationCreate,
    db: Session = Depends(get_db)
):
    """学员报名参加考试"""
    # 验证考试存在
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    # 检查是否已报名
    existing = db.query(ExamRegistration).filter(
        ExamRegistration.student_id == registration.student_id,
        ExamRegistration.exam_id == exam_id,
        ExamRegistration.status != "cancelled"
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该学员已报名此考试")

    # 检查人数限制
    if exam.max_participants:
        count = db.query(func.count(ExamRegistration.id)).filter(
            ExamRegistration.exam_id == exam_id,
            ExamRegistration.status != "cancelled"
        ).scalar()
        if count >= exam.max_participants:
            raise HTTPException(status_code=400, detail="考试报名人数已满")

    reg = ExamRegistration(
        student_id=registration.student_id,
        exam_id=exam_id,
        seat_number=registration.seat_number,
    )
    db.add(reg)
    db.commit()
    db.refresh(reg)
    return reg


@router.get("/{exam_id}/registrations", summary="获取考试报名列表")
def get_exam_registrations(
    exam_id: int,
    db: Session = Depends(get_db)
):
    """获取某次考试的报名列表"""
    registrations = db.query(ExamRegistration).filter(
        ExamRegistration.exam_id == exam_id
    ).all()

    items = []
    for r in registrations:
        student = r.student
        items.append({
            "id": r.id,
            "student_id": r.student_id,
            "student_name": student.name if student else None,
            "exam_id": r.exam_id,
            "registration_time": str(r.registration_time),
            "status": r.status,
            "seat_number": r.seat_number,
        })

    return {"code": 200, "message": "success", "data": items}


@router.put("/registrations/{registration_id}", summary="更新报名状态")
def update_registration(
    registration_id: int,
    update_data: ExamRegistrationUpdate,
    db: Session = Depends(get_db)
):
    """更新报名状态（签到、取消等）"""
    reg = db.query(ExamRegistration).filter(ExamRegistration.id == registration_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="报名记录不存在")

    if update_data.status:
        reg.status = update_data.status
    if update_data.seat_number:
        reg.seat_number = update_data.seat_number

    db.commit()
    db.refresh(reg)
    return {"code": 200, "message": "更新成功", "data": {
        "id": reg.id,
        "status": reg.status,
        "seat_number": reg.seat_number,
    }}


# ==================== 数据导出 ====================

# Excel 样式常量
_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

_REG_STATUS_MAP = {
    "registered": "已报名", "checked_in": "已签到",
    "absent": "缺考", "cancelled": "已取消"
}


def _style_header(ws, col_count):
    """为表头行添加样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    # 自适应列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)


@router.get("/export/participants", summary="导出考试人员")
def export_participants(
    exam_id: int = Query(..., description="考试ID"),
    db: Session = Depends(get_db)
):
    """导出指定考试的报名人员列表为 Excel"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    registrations = db.query(ExamRegistration).filter(
        ExamRegistration.exam_id == exam_id
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "考试人员"

    # 表头
    headers = ["序号", "学员姓名", "性别", "联系电话", "学校/单位", "专业", "报名时间", "签到状态", "座位号"]
    ws.append(headers)

    for idx, reg in enumerate(registrations, 1):
        student = reg.student
        ws.append([
            idx,
            student.name if student else "-",
            student.gender if student else "-",
            student.phone if student else "-",
            student.school if student else "-",
            student.major if student else "-",
            str(reg.registration_time) if reg.registration_time else "-",
            _REG_STATUS_MAP.get(reg.status, reg.status),
            reg.seat_number or "-",
        ])

    _style_header(ws, len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"考试人员_{exam.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/export/scores", summary="导出考试成绩")
def export_scores(
    exam_id: int = Query(..., description="考试ID"),
    db: Session = Depends(get_db)
):
    """导出指定考试的成绩列表为 Excel"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    scores = db.query(Score).filter(Score.exam_id == exam_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "考试成绩"

    headers = ["序号", "学员姓名", "性别", "学校/单位", "得分", "是否及格", "排名", "备注"]
    ws.append(headers)

    for idx, s in enumerate(scores, 1):
        student = s.student
        ws.append([
            idx,
            student.name if student else "-",
            student.gender if student else "-",
            student.school if student else "-",
            s.score,
            "是" if s.passed else "否",
            s.rank or "-",
            s.remarks or "-",
        ])

    _style_header(ws, len(headers))

    # 添加统计行
    if scores:
        ws.append([])
        avg = sum(s.score for s in scores) / len(scores)
        passed = sum(1 for s in scores if s.passed)
        ws.append(["", f"统计：共{len(scores)}人，平均分{avg:.1f}，及格{passed}人，及格率{passed/len(scores)*100:.1f}%"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"考试成绩_{exam.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/export/all", summary="导出考试数据（人员+成绩）")
def export_all(
    exam_id: int = Query(..., description="考试ID"),
    db: Session = Depends(get_db)
):
    """导出指定考试的人员和成绩到一个 Excel 文件（两个工作表）"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    wb = Workbook()

    # 工作表1：考试人员
    ws1 = wb.active
    ws1.title = "考试人员"
    headers1 = ["序号", "学员姓名", "性别", "联系电话", "学校/单位", "专业", "报名时间", "签到状态", "座位号"]
    ws1.append(headers1)

    registrations = db.query(ExamRegistration).filter(
        ExamRegistration.exam_id == exam_id
    ).all()

    for idx, reg in enumerate(registrations, 1):
        student = reg.student
        ws1.append([
            idx,
            student.name if student else "-",
            student.gender if student else "-",
            student.phone if student else "-",
            student.school if student else "-",
            student.major if student else "-",
            str(reg.registration_time) if reg.registration_time else "-",
            _REG_STATUS_MAP.get(reg.status, reg.status),
            reg.seat_number or "-",
        ])
    _style_header(ws1, len(headers1))

    # 工作表2：考试成绩
    ws2 = wb.create_sheet("考试成绩")
    headers2 = ["序号", "学员姓名", "性别", "学校/单位", "得分", "是否及格", "排名", "备注"]
    ws2.append(headers2)

    scores = db.query(Score).filter(Score.exam_id == exam_id).all()

    for idx, s in enumerate(scores, 1):
        student = s.student
        ws2.append([
            idx,
            student.name if student else "-",
            student.gender if student else "-",
            student.school if student else "-",
            s.score,
            "是" if s.passed else "否",
            s.rank or "-",
            s.remarks or "-",
        ])
    _style_header(ws2, len(headers2))

    if scores:
        ws2.append([])
        avg = sum(s.score for s in scores) / len(scores)
        passed = sum(1 for s in scores if s.passed)
        ws2.append(["", f"统计：共{len(scores)}人，平均分{avg:.1f}，及格{passed}人，及格率{passed/len(scores)*100:.1f}%"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"考试数据_{exam.name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )
